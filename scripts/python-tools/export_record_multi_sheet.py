# -*- coding: utf-8 -*-
"""
导出多个班级练习记录（每个班级一个 sheet，仅支持 xlsx）

对齐 Yii2 脚本：console/controllers/StudentController.php::actionExportRecordMultiSheet()

用法示例：
  python scripts/python-tools/export_record_multi_sheet.py "414,415,416" 2025-09-15 2025-12-31 "25秋季所有北外国商" xlsx
"""

from __future__ import annotations

import argparse
import os
import re
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from pymysql.err import OperationalError

from config import OUTPUT_DIR
from db import get_connection


EXPORT_TITLE: List[str] = [
    "姓名",
    "学号",
    "手机号",
    "是否是老师",
    "总次数",
    "总时长",
    "听力专项提升练习次数",
    "听力专项提升练习时长",
    "听力专项提升练习平均正确率",
    "阅读专项提升练习次数",
    "阅读专项提升练习时长",
    "阅读专项提升练习平均正确率",
    "写作专项提升练习次数",
    "写作专项提升练习时长",
    "写作专项提升练习平均正确率",
    "听力真题练习次数",
    "听力真题练习时长",
    "听力真题练习平均正确率",
    "阅读真题练习次数",
    "阅读真题练习时长",
    "阅读真题练习平均正确率",
    "大作文写作次数",
    "小作文写作次数",
    "大作文范文次数",
    "大作文练习次数",
    "口语机经练习次数",
    "口语机经练习时长",
    "口语进阶练习次数",
    "口语进阶练习时长",
    "模考次数",
    "模考时长",
]

FIELDS: List[str] = [
    "name",
    "account",
    "mobile",
    "is_teacher",
    "total_num",
    "total_time",
    "listening_special_improve_num",
    "listening_special_improve_time",
    "listening_special_improve_rate",
    "reading_special_improve_num",
    "reading_special_improve_time",
    "reading_special_improve_rate",
    "writing_special_improve_num",
    "writing_special_improve_time",
    "writing_special_improve_rate",
    "listening_real_num",
    "listening_real_time",
    "listening_real_rate",
    "reading_real_num",
    "reading_real_time",
    "reading_real_rate",
    "big_essay_num",
    "small_essay_num",
    "big_essay_model_num",
    "big_essay_exam_num",
    "oral_num",
    "oral_time",
    "oral_advanced_num",
    "oral_advanced_time",
    "mock_num",
    "mock_time",
]


def parse_class_ids(value: str) -> List[int]:
    ids = [int(x) for x in re.split(r"[,\s]+", value.strip()) if x.strip().isdigit()]
    ids = [x for x in ids if x > 0]
    seen = set()
    unique_ids: List[int] = []
    for x in ids:
        if x in seen:
            continue
        seen.add(x)
        unique_ids.append(x)
    return unique_ids


def parse_date_to_timestamp(date_str: str) -> int:
    dt = datetime.strptime(date_str, "%Y-%m-%d")
    return int(time.mktime(dt.timetuple()))


def sanitize_file_base_name(file_name: str) -> str:
    name = file_name.strip().replace("/", "_").replace("\\", "_")
    if not name:
        return ""
    base = Path(name).stem
    return base or name


def build_row_template(name: str = "", account: str = "", mobile: str = "") -> Dict[str, Any]:
    return {
        "name": name,
        "account": account,
        "mobile": mobile,
        "is_teacher": "否",
        "total_num": 0,
        "total_time": 0,
        "listening_special_improve_num": 0,
        "listening_special_improve_time": 0,
        "listening_special_improve_rate": 0,
        "reading_special_improve_num": 0,
        "reading_special_improve_time": 0,
        "reading_special_improve_rate": 0,
        "writing_special_improve_num": 0,
        "writing_special_improve_time": 0,
        "writing_special_improve_rate": 0,
        "listening_real_num": 0,
        "listening_real_time": 0,
        "listening_real_rate": 0,
        "reading_real_num": 0,
        "reading_real_time": 0,
        "reading_real_rate": 0,
        "big_essay_num": 0,
        "small_essay_num": 0,
        "big_essay_model_num": 0,
        "big_essay_exam_num": 0,
        "oral_num": 0,
        "oral_time": 0,
        "oral_advanced_num": 0,
        "oral_advanced_time": 0,
        "mock_num": 0,
        "mock_time": 0,
    }


def to_row_values(row: Dict[str, Any]) -> List[Any]:
    return [row.get(field) for field in FIELDS]


def sql_in_placeholders(values: Sequence[Any]) -> str:
    return ",".join(["%s"] * len(values))


def calculate_used_seconds_from_surplus_time(surplus_time: Optional[int], total_seconds: int) -> int:
    if total_seconds <= 0 or surplus_time is None:
        return 0
    surplus_time = int(surplus_time)
    if surplus_time <= 0:
        return total_seconds
    if surplus_time >= total_seconds:
        return 0
    return total_seconds - surplus_time


def calculate_mock_duration_from_parts(
    listening_surplus: Optional[int], reading_surplus: Optional[int], writing_surplus: Optional[int]
) -> int:
    duration = 0
    duration += calculate_used_seconds_from_surplus_time(listening_surplus, 1800)
    duration += calculate_used_seconds_from_surplus_time(reading_surplus, 3600)
    duration += calculate_used_seconds_from_surplus_time(writing_surplus, 3600)
    return duration


def make_unique_sheet_title(raw_title: str, used_titles: Dict[str, bool]) -> str:
    title = raw_title.strip()
    title = re.sub(r"[\\/?*\[\]:]", " ", title)
    title = re.sub(r"\s+", " ", title).strip()
    if not title:
        title = "Sheet"

    title = title[:31]
    candidate = title
    suffix_index = 2
    while candidate in used_titles:
        suffix = f"-{suffix_index}"
        allowed_len = 31 - len(suffix)
        candidate = title[:allowed_len] + suffix
        suffix_index += 1

    used_titles[candidate] = True
    return candidate


def set_auto_width(sheet) -> None:
    dims: Dict[int, int] = {}
    for row in sheet.iter_rows(values_only=True):
        for idx, value in enumerate(row, start=1):
            if value is None:
                length = 0
            else:
                length = len(str(value))
            dims[idx] = max(dims.get(idx, 0), length)
    for idx, length in dims.items():
        sheet.column_dimensions[get_column_letter(idx)].width = min(max(length + 2, 10), 60)


@dataclass
class ExportResult:
    title: List[str]
    rows: List[List[Any]]
    data: Dict[int, Dict[str, Any]]
    teacher_user_ids: List[int]


def build_export_record_rows_for_class_ids(
    cursor, class_ids: List[int], start_time: int, end_time: int, append_summary_rows: bool
) -> ExportResult:
    title = list(EXPORT_TITLE)
    data: Dict[int, Dict[str, Any]] = {}

    # 学生列表
    placeholders = sql_in_placeholders(class_ids)
    cursor.execute(
        f"SELECT student_id, student_name FROM edu_class_student WHERE class_id IN ({placeholders})",
        tuple(class_ids),
    )
    for row in cursor.fetchall():
        student_id = int(row["student_id"])
        if student_id not in data:
            data[student_id] = build_row_template(name=row.get("student_name") or "")

    # 老师列表
    cursor.execute(
        f"SELECT teacher_id FROM edu_class_teacher WHERE class_id IN ({placeholders})",
        tuple(class_ids),
    )
    teacher_ids = [int(x["teacher_id"]) for x in cursor.fetchall() if x.get("teacher_id") is not None]
    teacher_user_ids: List[int] = []
    if teacher_ids:
        teacher_placeholders = sql_in_placeholders(teacher_ids)
        cursor.execute(
            f"SELECT id, user_id, name FROM edu_teacher WHERE id IN ({teacher_placeholders})",
            tuple(teacher_ids),
        )
        for teacher in cursor.fetchall():
            user_id = int(teacher["user_id"])
            teacher_user_ids.append(user_id)
            if user_id not in data:
                data[user_id] = build_row_template(name=teacher.get("name") or "")
            data[user_id]["is_teacher"] = "是"

    user_ids = list(data.keys())
    if not user_ids:
        return ExportResult(title=title, rows=[], data={}, teacher_user_ids=[])

    # 用户信息（学号、手机号）
    user_placeholders = sql_in_placeholders(user_ids)
    cursor.execute(
        f"SELECT id, mobile, account FROM student WHERE id IN ({user_placeholders})",
        tuple(user_ids),
    )
    for user in cursor.fetchall():
        uid = int(user["id"])
        if uid not in data:
            continue
        data[uid]["mobile"] = user.get("mobile") or ""
        data[uid]["account"] = user.get("account") or ""

    writing_special: Dict[int, List[float]] = {}
    reading_special: Dict[int, List[float]] = {}
    listening_special: Dict[int, List[float]] = {}
    listening_real: Dict[int, List[float]] = {}
    reading_real: Dict[int, List[float]] = {}

    # 专项提升：exam_collection_record + exam_question_collection(type)
    try:
        cursor.execute(
            f"""
            SELECT student_id, collection_id, duration, rate
            FROM exam_collection_record
            WHERE student_id IN ({user_placeholders})
              AND status = 2
              AND update_time >= %s AND update_time <= %s
            """,
            tuple(user_ids) + (start_time, end_time),
        )
        collection_records = cursor.fetchall()
        collection_record_has_duration = True
    except OperationalError:
        cursor.execute(
            f"""
            SELECT student_id, collection_id, start_time, end_time, rate
            FROM exam_collection_record
            WHERE student_id IN ({user_placeholders})
              AND status = 2
              AND update_time >= %s AND update_time <= %s
            """,
            tuple(user_ids) + (start_time, end_time),
        )
        collection_records = cursor.fetchall()
        collection_record_has_duration = False

    if collection_records:
        collection_ids = list({int(r["collection_id"]) for r in collection_records if r.get("collection_id") is not None})
        collection_type_map: Dict[int, int] = {}
        if collection_ids:
            col_placeholders = sql_in_placeholders(collection_ids)
            cursor.execute(
                f"SELECT id, type FROM exam_question_collection WHERE id IN ({col_placeholders})",
                tuple(collection_ids),
            )
            for col in cursor.fetchall():
                collection_type_map[int(col["id"])] = int(col["type"])

        for record in collection_records:
            sid = int(record["student_id"])
            if sid not in data:
                continue
            cid = record.get("collection_id")
            if cid is None:
                continue
            ctype = collection_type_map.get(int(cid))
            if ctype is None:
                continue

            if collection_record_has_duration:
                duration = int(record.get("duration") or 0)
            else:
                start = int(record.get("start_time") or 0)
                end = int(record.get("end_time") or 0)
                duration = max(end - start, 0)

            rate_raw = record.get("rate")
            try:
                rate = float(rate_raw) if rate_raw is not None else 0.0
            except (TypeError, ValueError):
                rate = 0.0

            data[sid]["total_num"] += 1
            data[sid]["total_time"] += duration
            if ctype == 1:
                data[sid]["writing_special_improve_num"] += 1
                data[sid]["writing_special_improve_time"] += duration
                writing_special.setdefault(sid, []).append(rate)
            elif ctype == 2:
                data[sid]["reading_special_improve_num"] += 1
                data[sid]["reading_special_improve_time"] += duration
                reading_special.setdefault(sid, []).append(rate)
            elif ctype == 3:
                data[sid]["listening_special_improve_num"] += 1
                data[sid]["listening_special_improve_time"] += duration
                listening_special.setdefault(sid, []).append(rate)

    for sid, rates in writing_special.items():
        if rates:
            data[sid]["writing_special_improve_rate"] = round(sum(rates) / len(rates), 4)
    for sid, rates in reading_special.items():
        if rates:
            data[sid]["reading_special_improve_rate"] = round(sum(rates) / len(rates), 4)
    for sid, rates in listening_special.items():
        if rates:
            data[sid]["listening_special_improve_rate"] = round(sum(rates) / len(rates), 4)

    # 听力真题
    cursor.execute(
        f"""
        SELECT student_id, duration, correct, total
        FROM listening_exam_record
        WHERE student_id IN ({user_placeholders})
          AND status = 2
          AND update_time >= %s AND update_time <= %s
        """,
        tuple(user_ids) + (start_time, end_time),
    )
    for record in cursor.fetchall():
        sid = int(record["student_id"])
        if sid not in data:
            continue
        duration = int(record.get("duration") or 0)
        correct = int(record.get("correct") or 0)
        total = int(record.get("total") or 0)
        data[sid]["total_num"] += 1
        data[sid]["total_time"] += duration
        data[sid]["listening_real_num"] += 1
        data[sid]["listening_real_time"] += duration
        if total > 0:
            listening_real.setdefault(sid, []).append(correct / total)

    for sid, rates in listening_real.items():
        if rates:
            data[sid]["listening_real_rate"] = round(sum(rates) / len(rates), 4)

    # 阅读真题
    cursor.execute(
        f"""
        SELECT student_id, duration, correct, total
        FROM reading_exam_record
        WHERE student_id IN ({user_placeholders})
          AND status = 2
          AND finished_time >= %s AND finished_time <= %s
        """,
        tuple(user_ids) + (start_time, end_time),
    )
    for record in cursor.fetchall():
        sid = int(record["student_id"])
        if sid not in data:
            continue
        duration = int(record.get("duration") or 0)
        correct = int(record.get("correct") or 0)
        total = int(record.get("total") or 0)
        data[sid]["total_num"] += 1
        data[sid]["total_time"] += duration
        data[sid]["reading_real_num"] += 1
        data[sid]["reading_real_time"] += duration
        if total > 0:
            reading_real.setdefault(sid, []).append(correct / total)

    for sid, rates in reading_real.items():
        if rates:
            data[sid]["reading_real_rate"] = round(sum(rates) / len(rates), 4)

    # 大作文写作
    cursor.execute(
        f"""
        SELECT student_id
        FROM writing_big_essay_record
        WHERE student_id IN ({user_placeholders})
          AND status = 2
          AND update_time >= %s AND update_time <= %s
        """,
        tuple(user_ids) + (start_time, end_time),
    )
    for record in cursor.fetchall():
        sid = int(record["student_id"])
        if sid in data:
            data[sid]["big_essay_num"] += 1
            data[sid]["total_num"] += 1

    # 小作文写作
    cursor.execute(
        f"""
        SELECT student_id
        FROM writing_essay_record
        WHERE student_id IN ({user_placeholders})
          AND status = 2
          AND update_time >= %s AND update_time <= %s
        """,
        tuple(user_ids) + (start_time, end_time),
    )
    for record in cursor.fetchall():
        sid = int(record["student_id"])
        if sid in data:
            data[sid]["small_essay_num"] += 1
            data[sid]["total_num"] += 1

    # 大作文范文
    cursor.execute(
        f"""
        SELECT student_id
        FROM writing_big_essay_sample_text
        WHERE student_id IN ({user_placeholders})
          AND status = 2
          AND update_time >= %s AND update_time <= %s
        """,
        tuple(user_ids) + (start_time, end_time),
    )
    for record in cursor.fetchall():
        sid = int(record["student_id"])
        if sid in data:
            data[sid]["big_essay_model_num"] += 1
            data[sid]["total_num"] += 1

    # 大作文练习（不筛 status，对齐 PHP）
    cursor.execute(
        f"""
        SELECT student_id
        FROM writing_practice_record
        WHERE student_id IN ({user_placeholders})
          AND update_time >= %s AND update_time <= %s
        """,
        tuple(user_ids) + (start_time, end_time),
    )
    for record in cursor.fetchall():
        sid = int(record["student_id"])
        if sid in data:
            data[sid]["big_essay_exam_num"] += 1
            data[sid]["total_num"] += 1

    # 口语机经练习（duration 毫秒 -> 秒）
    cursor.execute(
        f"""
        SELECT student_id, duration
        FROM speaking_exam_dialogue_log
        WHERE student_id IN ({user_placeholders})
          AND role = 1
          AND type IN (1, 2)
          AND update_time >= %s AND update_time <= %s
        """,
        tuple(user_ids) + (start_time, end_time),
    )
    for record in cursor.fetchall():
        sid = int(record["student_id"])
        if sid not in data:
            continue
        duration = (record.get("duration") or 0) / 1000
        data[sid]["oral_num"] += 1
        data[sid]["oral_time"] += duration
        data[sid]["total_num"] += 1
        data[sid]["total_time"] += duration

    # 口语进阶练习（duration 毫秒 -> 秒）
    cursor.execute(
        f"""
        SELECT student_id, duration
        FROM speaking_advance_record
        WHERE student_id IN ({user_placeholders})
          AND update_time >= %s AND update_time <= %s
        """,
        tuple(user_ids) + (start_time, end_time),
    )
    for record in cursor.fetchall():
        sid = int(record["student_id"])
        if sid not in data:
            continue
        duration = (record.get("duration") or 0) / 1000
        data[sid]["oral_advanced_num"] += 1
        data[sid]["oral_advanced_time"] += duration
        data[sid]["total_num"] += 1
        data[sid]["total_time"] += duration

    # 模考：simulate_exam_record + parts(surplus_time)
    cursor.execute(
        f"""
        SELECT id, student_id
        FROM simulate_exam_record
        WHERE student_id IN ({user_placeholders})
          AND update_time >= %s AND update_time <= %s
        """,
        tuple(user_ids) + (start_time, end_time),
    )
    mock_records = cursor.fetchall()
    if mock_records:
        record_ids = [int(r["id"]) for r in mock_records]
        record_placeholders = sql_in_placeholders(record_ids)

        cursor.execute(
            f"SELECT record_id, surplus_time FROM simulate_exam_listening WHERE record_id IN ({record_placeholders})",
            tuple(record_ids),
        )
        listening_by_record = {int(r["record_id"]): r.get("surplus_time") for r in cursor.fetchall()}

        cursor.execute(
            f"SELECT record_id, surplus_time FROM simulate_exam_reading WHERE record_id IN ({record_placeholders})",
            tuple(record_ids),
        )
        reading_by_record = {int(r["record_id"]): r.get("surplus_time") for r in cursor.fetchall()}

        cursor.execute(
            f"SELECT record_id, surplus_time FROM simulate_exam_writing WHERE record_id IN ({record_placeholders})",
            tuple(record_ids),
        )
        writing_by_record = {int(r["record_id"]): r.get("surplus_time") for r in cursor.fetchall()}

        for record in mock_records:
            sid = int(record["student_id"])
            if sid not in data:
                continue
            rid = int(record["id"])
            duration = calculate_mock_duration_from_parts(
                listening_by_record.get(rid), reading_by_record.get(rid), writing_by_record.get(rid)
            )
            data[sid]["mock_num"] += 1
            data[sid]["mock_time"] += duration
            data[sid]["total_num"] += 1
            data[sid]["total_time"] += duration

    rows = [to_row_values(row) for row in data.values()]

    if append_summary_rows:
        total_practice_num = 0
        total_practice_time = 0
        practiced_student_count = 0
        overall_rate_pool: List[float] = []

        count_fields = [
            "listening_special_improve_num",
            "reading_special_improve_num",
            "writing_special_improve_num",
            "listening_real_num",
            "reading_real_num",
            "big_essay_num",
            "small_essay_num",
            "big_essay_model_num",
            "big_essay_exam_num",
            "oral_num",
            "oral_advanced_num",
            "mock_num",
        ]
        time_fields = [
            "listening_special_improve_time",
            "reading_special_improve_time",
            "writing_special_improve_time",
            "listening_real_time",
            "reading_real_time",
            "oral_time",
            "oral_advanced_time",
            "mock_time",
        ]
        rate_fields = {
            "listening_special_improve_rate": "listening_special_improve_num",
            "reading_special_improve_rate": "reading_special_improve_num",
            "writing_special_improve_rate": "writing_special_improve_num",
            "listening_real_rate": "listening_real_num",
            "reading_real_rate": "reading_real_num",
        }

        count_sums = {k: 0 for k in count_fields}
        time_sums = {k: 0 for k in time_fields}
        rate_sums: Dict[str, float] = {}
        rate_counts: Dict[str, int] = {}

        teacher_user_id_set = set(teacher_user_ids)
        for user_id, row in data.items():
            if user_id in teacher_user_id_set:
                continue
            total_practice_num += row["total_num"]
            total_practice_time += row["total_time"]
            if row["total_num"] > 0:
                practiced_student_count += 1
            for field in count_fields:
                count_sums[field] += row[field]
            for field in time_fields:
                time_sums[field] += row[field]
            for rate_field, count_field in rate_fields.items():
                if row[count_field] > 0 and row[rate_field] > 0:
                    rate_sums[rate_field] = rate_sums.get(rate_field, 0.0) + float(row[rate_field])
                    rate_counts[rate_field] = rate_counts.get(rate_field, 0) + 1
                    overall_rate_pool.append(float(row[rate_field]))

        avg_practice_duration = round(total_practice_time / practiced_student_count, 2) if practiced_student_count > 0 else 0
        overall_avg_accuracy = round(sum(overall_rate_pool) / len(overall_rate_pool), 4) if overall_rate_pool else 0
        rate_averages = {
            rate_field: round(rate_sums[rate_field] / rate_counts[rate_field], 4)
            if rate_counts.get(rate_field, 0) > 0
            else 0
            for rate_field in rate_fields.keys()
        }

        summary_title_row = build_row_template("标题")
        summary_title_row["account"] = "练习人数"
        summary_title_row["mobile"] = "练过的人的平均练习时长(秒)"
        summary_title_row["name"] = "练过的人的平均正确率"
        summary_title_row["total_num"] = "练习总次数"
        summary_title_row["total_time"] = "练习总时长(秒)"
        summary_title_row["listening_special_improve_num"] = "听力专项次数总数"
        summary_title_row["listening_special_improve_time"] = "听力专项时长总计(秒)"
        summary_title_row["listening_special_improve_rate"] = "听力专项平均正确率"
        summary_title_row["reading_special_improve_num"] = "阅读专项次数总数"
        summary_title_row["reading_special_improve_time"] = "阅读专项时长总计(秒)"
        summary_title_row["reading_special_improve_rate"] = "阅读专项平均正确率"
        summary_title_row["writing_special_improve_num"] = "写作专项次数总数"
        summary_title_row["writing_special_improve_time"] = "写作专项时长总计(秒)"
        summary_title_row["writing_special_improve_rate"] = "写作专项平均正确率"
        summary_title_row["listening_real_num"] = "听力真题次数总数"
        summary_title_row["listening_real_time"] = "听力真题时长总计(秒)"
        summary_title_row["listening_real_rate"] = "听力真题平均正确率"
        summary_title_row["reading_real_num"] = "阅读真题次数总数"
        summary_title_row["reading_real_time"] = "阅读真题时长总计(秒)"
        summary_title_row["reading_real_rate"] = "阅读真题平均正确率"
        summary_title_row["big_essay_num"] = "大作文写作次数总数"
        summary_title_row["small_essay_num"] = "小作文写作次数总数"
        summary_title_row["big_essay_model_num"] = "大作文范文次数总数"
        summary_title_row["big_essay_exam_num"] = "大作文练习次数总数"
        summary_title_row["oral_num"] = "口语机经练习次数总数"
        summary_title_row["oral_time"] = "口语机经练习时长总计(秒)"
        summary_title_row["oral_advanced_num"] = "口语进阶练习次数总数"
        summary_title_row["oral_advanced_time"] = "口语进阶练习时长总计(秒)"
        summary_title_row["mock_num"] = "模考次数总数"
        summary_title_row["mock_time"] = "模考时长总计(秒)"

        summary_value_row = build_row_template("统计")
        summary_value_row["name"] = overall_avg_accuracy
        summary_value_row["account"] = practiced_student_count
        summary_value_row["mobile"] = avg_practice_duration
        summary_value_row["total_num"] = total_practice_num
        summary_value_row["total_time"] = total_practice_time
        for field in count_fields:
            summary_value_row[field] = count_sums[field]
        for field in time_fields:
            summary_value_row[field] = time_sums[field]
        for rate_field in rate_fields.keys():
            summary_value_row[rate_field] = rate_averages.get(rate_field, 0)

        rows.append(to_row_values(summary_title_row))
        rows.append(to_row_values(summary_value_row))

    return ExportResult(title=title, rows=rows, data=data, teacher_user_ids=teacher_user_ids)


def fetch_class_name_map(cursor, class_ids: List[int]) -> Dict[int, str]:
    placeholders = sql_in_placeholders(class_ids)
    cursor.execute(f"SELECT id, name FROM edu_class WHERE id IN ({placeholders})", tuple(class_ids))
    return {int(row["id"]): (row.get("name") or "") for row in cursor.fetchall()}


def export_multi_sheet(
    class_ids: List[int],
    start_time: int,
    end_time: int,
    file_base_name: str,
    output_dir: Optional[os.PathLike | str] = None,
) -> str:
    output_dir = Path(output_dir) if output_dir is not None else Path(OUTPUT_DIR)
    output_dir.mkdir(parents=True, exist_ok=True)
    file_path = output_dir / f"{file_base_name}.xlsx"

    wb = Workbook()
    used_titles: Dict[str, bool] = {}

    conn = get_connection()
    try:
        with conn.cursor() as cursor:
            class_name_map = fetch_class_name_map(cursor, class_ids)

            for sheet_index, class_id in enumerate(class_ids):
                class_name = class_name_map.get(class_id) or f"班级{class_id}"
                raw_title = f"{class_name}({class_id})"
                sheet_title = make_unique_sheet_title(raw_title, used_titles)

                if sheet_index == 0:
                    sheet = wb.active
                else:
                    sheet = wb.create_sheet(index=sheet_index)
                sheet.title = sheet_title

                export_result = build_export_record_rows_for_class_ids(
                    cursor=cursor,
                    class_ids=[class_id],
                    start_time=start_time,
                    end_time=end_time,
                    append_summary_rows=True,
                )

                sheet.append(export_result.title)
                for row in export_result.rows:
                    sheet.append(row)
                set_auto_width(sheet)
    finally:
        conn.close()

    wb.save(str(file_path))
    return str(file_path)


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="导出多个班级学生练习记录（每班一个 sheet，仅支持 xlsx）")
    parser.add_argument("class_ids", help="逗号或空格分隔的班级ID，例如: 414,415,416")
    parser.add_argument("start_date", help="开始日期，格式: YYYY-MM-DD")
    parser.add_argument("end_date", help="结束日期，格式: YYYY-MM-DD")
    parser.add_argument("file_name", help="导出文件名（不含扩展名）")
    parser.add_argument("format", nargs="?", default="xlsx", help="导出格式，仅支持 xlsx（兼容原命令参数）")
    return parser


def main(argv: Optional[Sequence[str]] = None) -> int:
    load_dotenv()
    parser = build_arg_parser()
    args = parser.parse_args(argv)

    class_ids = parse_class_ids(args.class_ids)
    if not class_ids:
        print("class_ids 不能为空")
        return 1

    raw_file_name = (args.file_name or "").strip().replace("/", "_").replace("\\", "_")
    if raw_file_name == "":
        print("导出文件名称不能为空")
        return 1
    file_base_name = sanitize_file_base_name(raw_file_name)
    if file_base_name == "":
        file_base_name = f"export_{time.strftime('%Y%m%d_%H%M%S')}"

    fmt = (args.format or "xlsx").lower().strip()
    if fmt != "xlsx":
        print("格式错误，仅支持 xlsx")
        return 1

    try:
        start_time = parse_date_to_timestamp(args.start_date)
        end_time = parse_date_to_timestamp(args.end_date)
    except ValueError:
        print("日期格式错误，请使用 Y-m-d 格式，例如: 2025-05-12")
        return 1
    if start_time > end_time:
        print("开始日期不能大于结束日期")
        return 1

    file_path = export_multi_sheet(class_ids, start_time, end_time, file_base_name)
    print(f"导出成功，文件路径：{file_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
